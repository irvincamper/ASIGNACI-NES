from __future__ import annotations

from datetime import datetime
from pathlib import Path
from textwrap import wrap

from openpyxl import load_workbook
from reportlab.lib.colors import HexColor, black, white
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from utils.constants import DATA_DIR, EXPEDIENTES_DIR, IMPORTAR_DIR


RECURSOS_POR_PAGINA_FORMATO = 16


class PdfExpedienteService:
    BLUE = HexColor("#1782B5")
    BORDER = black

    def __init__(self) -> None:
        self.font = self._register_font()

    def generar_pdf_expediente(self, json_expediente: dict) -> str:
        recursos = list(json_expediente.get("recursos") or [])
        if not recursos:
            raise ValueError("El empleado no tiene asignaciones para generar el expediente.")

        EXPEDIENTES_DIR.mkdir(parents=True, exist_ok=True)
        matricula = str(json_expediente.get("empleado", {}).get("matricula", "")).strip()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = EXPEDIENTES_DIR / f"ALA-RH-FR-29_{matricula}_{stamp}.pdf"

        doc = canvas.Canvas(str(path), pagesize=letter)
        total_pages = max(2, (len(recursos) + RECURSOS_POR_PAGINA_FORMATO - 1) // RECURSOS_POR_PAGINA_FORMATO)
        for page_index in range(total_pages):
            offset = page_index * RECURSOS_POR_PAGINA_FORMATO
            chunk = recursos[offset : offset + RECURSOS_POR_PAGINA_FORMATO]
            self._draw_page(doc, json_expediente, chunk, page_index + 1, total_pages, offset)
            doc.showPage()
        doc.save()
        return str(path)

    def _draw_page(self, doc: canvas.Canvas, data: dict, recursos: list[dict], page_number: int, total_pages: int, offset: int) -> None:
        page_width, page_height = letter
        margin = 16
        table_width = page_width - margin * 2
        y = page_height - margin
        logo_path = self._template_logo_path()

        self._cell(
            doc,
            margin,
            y - 40,
            table_width,
            40,
            "Entrega-recepción de equipos, herramientas, uniformes, EPP y papelería",
            bold=True,
            font_size=12,
        )
        if logo_path:
            try:
                doc.drawImage(ImageReader(str(logo_path)), margin + 5, y - 39, width=34, height=48, preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        y -= 50

        empleado = data.get("empleado", {})
        self._cell(doc, margin, y - 28, 130, 28, "Nombre del trabajador:", fill=self.BLUE, text_color=white, bold=True)
        self._cell(doc, margin + 130, y - 28, table_width - 250, 28, str(empleado.get("nombre", "")), font_size=9)
        self._cell(doc, margin + table_width - 120, y - 28, 70, 28, "Matrícula:", fill=self.BLUE, text_color=white, bold=True)
        self._cell(doc, margin + table_width - 50, y - 28, 50, 28, str(empleado.get("matricula", "")), font_size=9)
        y -= 38

        self._text(doc, margin, y, "Los motivos de entrega-recepción pueden ser: CONTRATACIÓN, CAMBIO DE PUESTO O BAJA.", 8)
        y -= 10

        reason_widths = [220, table_width - 340, 120]
        x = margin
        for label, width in zip(["Motivo de entrega - recepción", "Puesto", "Fecha"], reason_widths):
            self._cell(doc, x, y - 16, width, 16, label, fill=self.BLUE, text_color=white, bold=True, font_size=8)
            x += width
        y -= 16
        motivo = str(data.get("motivo", "Contratación"))
        for option in ["Contratación", "Cambio de puesto", "Baja"]:
            x = margin
            mark = "X" if option == motivo else ""
            self._cell(doc, x, y - 16, reason_widths[0], 16, f"{mark}   {option}", align="left", font_size=8)
            x += reason_widths[0]
            self._cell(doc, x, y - 16, reason_widths[1], 16, str(data.get("puesto", "")), font_size=7)
            x += reason_widths[1]
            self._cell(doc, x, y - 16, reason_widths[2], 16, str(data.get("fecha_documento", "")), font_size=8)
            y -= 16
        y -= 10

        self._cell(doc, margin, y - 18, table_width, 18, "RECURSOS PARA ASIGNACIÓN Y/O DEVOLUCIÓN", bold=True, font_size=9, bottom_only=True)
        y -= 24

        col_weights = [3.7, 3.2, 13, 12.7, 3.2, 13, 13, 13, 13, 13, 33.6, 8.6, 13, 15.7]
        total_weight = sum(col_weights)
        widths = [table_width * weight / total_weight for weight in col_weights]
        headers = ["No.", "Asignación", "Devolución", "Fecha", "Equipo", "Herramienta", "Uniforme", "EPP", "Papelería", "Otros", "Descripción", "Cantidad", "Unidad", "Firma"]
        x = margin
        for header, width in zip(headers, widths):
            self._cell(doc, x, y - 30, width, 30, header, fill=self.BLUE, text_color=white, bold=True, font_size=6)
            x += width
        y -= 30

        row_height = 24
        for row_index in range(RECURSOS_POR_PAGINA_FORMATO):
            item = recursos[row_index] if row_index < len(recursos) else {}
            values = self._resource_values(item, offset + row_index + 1 if item else "")
            x = margin
            for column_index, (value, width) in enumerate(zip(values, widths)):
                align = "left" if column_index == 10 else "center"
                size = 6 if column_index != 10 else 6.5
                self._cell(doc, x, y - row_height, width, row_height, value, align=align, font_size=size)
                x += width
            y -= row_height

        y -= 4
        self._cell(doc, margin, y - 18, 90, 18, "Observaciones:", align="left", font_size=8, bottom_only=True)
        self._cell(doc, margin + 90, y - 18, table_width - 90, 18, str(data.get("observaciones", "")), align="left", font_size=8, bottom_only=True)
        y -= 28
        self._cell(doc, margin, y - 18, table_width, 18, "", bottom_only=True)
        y -= 18

        footer_left = "ALA-RH-FR-29 Entrega-recepción de equipos,\nherramientas, uniformes, EPP y papelería"
        self._text(doc, margin, y, footer_left, 7)
        self._text(doc, margin + table_width * 0.46, y - 1, f"Página {page_number} de {total_pages}", 7)
        self._text(doc, margin + table_width - 72, y - 1, "Revisión: 2", 7)

    def _resource_values(self, item: dict, number: int | str) -> list[str]:
        if not item:
            return [""] * 14
        operation = str(item.get("tipo_operacion", "Asignacion"))
        category_mark = self._category_mark(str(item.get("categoria", "")))
        category_values = ["" for _ in range(6)]
        category_values[category_mark] = "X"
        return [
            str(number),
            "X" if operation in {"Asignacion", "Asignación"} else "",
            "X" if operation in {"Devolucion", "Devolución"} else "",
            str(item.get("fecha") or ""),
            *category_values,
            str(item.get("objeto") or item.get("concepto") or ""),
            str(item.get("cantidad") or ""),
            str(item.get("unidad") or ""),
            "",
        ]

    def _category_mark(self, category: str) -> int:
        value = category.lower()
        if "herramient" in value:
            return 1
        if "uniform" in value or "vestimenta" in value:
            return 2
        if "seguridad" in value or "epp" in value:
            return 3
        if "papeler" in value:
            return 4
        if "equipo" in value or "tecnolog" in value:
            return 0
        return 5

    def _cell(
        self,
        doc: canvas.Canvas,
        x: float,
        y: float,
        width: float,
        height: float,
        text: str,
        *,
        fill=None,
        text_color=black,
        bold: bool = False,
        font_size: float = 8,
        align: str = "center",
        bottom_only: bool = False,
    ) -> None:
        if fill:
            doc.setFillColor(fill)
            doc.rect(x, y, width, height, fill=1, stroke=0)
        doc.setStrokeColor(self.BORDER)
        doc.setLineWidth(0.5)
        if bottom_only:
            doc.line(x, y, x + width, y)
        else:
            doc.rect(x, y, width, height, fill=0, stroke=1)
        if text:
            self._draw_wrapped_text(doc, x + 2, y + 2, width - 4, height - 4, str(text), text_color, bold, font_size, align)

    def _draw_wrapped_text(
        self,
        doc: canvas.Canvas,
        x: float,
        y: float,
        width: float,
        height: float,
        text: str,
        color,
        bold: bool,
        font_size: float,
        align: str,
    ) -> None:
        font_name = self.font
        doc.setFillColor(color)
        doc.setFont(font_name, font_size)
        approx_chars = max(1, int(width / max(font_size * 0.42, 1)))
        lines: list[str] = []
        for raw_line in text.splitlines() or [""]:
            lines.extend(wrap(raw_line, approx_chars) or [""])
        max_lines = max(1, int(height / (font_size + 1)))
        lines = lines[:max_lines]
        total_height = len(lines) * (font_size + 1)
        cursor_y = y + (height + total_height) / 2 - font_size
        for line in lines:
            if align == "left":
                cursor_x = x + 1
            else:
                cursor_x = x + (width - pdfmetrics.stringWidth(line, font_name, font_size)) / 2
            doc.drawString(cursor_x, cursor_y, line)
            cursor_y -= font_size + 1

    def _text(self, doc: canvas.Canvas, x: float, y: float, text: str, font_size: float) -> None:
        doc.setFont(self.font, font_size)
        doc.setFillColor(black)
        for index, line in enumerate(text.splitlines()):
            doc.drawString(x, y - index * (font_size + 2), line)

    def _template_logo_path(self) -> Path | None:
        output = DATA_DIR / "temp" / "formato_entrega_recepcion_logo.png"
        if output.exists():
            return output
        matches = list(IMPORTAR_DIR.glob("FORMATO ENTREGA-RECEPCI*.xlsx"))
        if not matches:
            return None
        try:
            workbook = load_workbook(matches[0])
            try:
                sheet = workbook["10"] if "10" in workbook.sheetnames else workbook[workbook.sheetnames[1]]
                if not getattr(sheet, "_images", None):
                    return None
                output.parent.mkdir(parents=True, exist_ok=True)
                output.write_bytes(sheet._images[0]._data())
                return output
            finally:
                workbook.close()
        except Exception:
            return None

    def _register_font(self) -> str:
        arial = Path("C:/Windows/Fonts/arial.ttf")
        if arial.exists():
            try:
                pdfmetrics.registerFont(TTFont("Arial", str(arial)))
                return "Arial"
            except Exception:
                pass
        return "Helvetica"
