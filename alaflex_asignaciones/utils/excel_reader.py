from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.normalizers import normalize_header


def open_workbook_readonly(path: Path):
    return load_workbook(path, read_only=True, data_only=True)


def get_sheet_by_name(workbook, expected_name: str):
    expected = normalize_header(expected_name)
    for sheet_name in workbook.sheetnames:
        if normalize_header(sheet_name) == expected:
            return workbook[sheet_name]
    return None


def find_header_row(
    worksheet: Worksheet,
    required_headers: Iterable[str],
    max_scan_rows: int = 20,
) -> tuple[int | None, list[str]]:
    required = [normalize_header(header) for header in required_headers]
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        headers = [normalize_header(value) for value in row]
        if all(header in headers for header in required):
            return row_index, headers
    return None, []


def headers_from_row(row: Iterable[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        header = normalize_header(value)
        if not header:
            header = f"columna {index}"
        if header in seen:
            seen[header] += 1
            header = f"{header} {seen[header]}"
        else:
            seen[header] = 1
        headers.append(header)
    return headers


def rows_as_dicts(worksheet: Worksheet, header_row: int) -> list[dict[str, Any]]:
    rows = worksheet.iter_rows(min_row=header_row, values_only=True)
    try:
        headers = headers_from_row(next(rows))
    except StopIteration:
        return []
    output: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=header_row + 1):
        item = {headers[index]: value for index, value in enumerate(row[: len(headers)])}
        item["_row_number"] = row_number
        output.append(item)
    return output
